"""Use case: ImportarRegistroProphet.

Lee un archivo Excel con el registro Prophet (formato de la ficha SMNYL)
y construye un Documento de tipo 'prophet' con sus secciones pre-pobladas
a partir de los datos encontrados en cada hoja.

Si se provee un LLMClient, usa Haiku 4.5 para la extracción estructurada.
Si llm=None, usa lógica heurística (útil para tests unitarios sin API key).
"""

from __future__ import annotations

import io
import json
from dataclasses import dataclass, field
from datetime import UTC, datetime
from typing import Any

import openpyxl
from anthropic.types import TextBlockParam

from src.core.models import Documento, EventoAuditoria
from src.core.models.documento import MetadataModelo
from src.core.template_catalog_prophet import (
    SeccionCatalogoProphet,
    construir_secciones_vacias_prophet,
    por_id_prophet,
)
from src.llm import LLMClient
from src.llm.prompts.extraer_seccion_prophet import (
    EXTRAER_SECCION_PROPHET_SYSTEM,
    construir_prompt_extraccion,
)
from src.storage.repositories import DocumentoRepository

# Mapea cada id de sección Prophet a los posibles nombres de hoja Excel
# (en minúsculas, sin acentos para tolerancia de variación de formato).
_HOJA_POR_SECCION: dict[str, list[str]] = {
    "identificacion": ["descripcion_general", "descripción_general", "modelos"],
    "objetivo_alcance": ["descripcion_general", "descripción_general"],
    "responsables_roles": ["descripcion_general", "descripción_general"],
    "corridas_runs": ["detalle runs", "detalle_runs", "runs", "corridas"],
    "variables_criticas": ["variables criticas", "variables_criticas", "variables"],
    "inputs_dependencias": ["descripcion_general", "descripción_general"],
    "matriz_conocimiento": ["conocimiento_tecnico", "conocimiento técnico", "conocimiento"],
}


@dataclass
class ResultadoImportacionProphet:
    documento: Documento | None
    secciones_importadas: int = 0
    secciones_vacias: int = 0
    advertencias: list[str] = field(default_factory=list)


class ImportarRegistroProphet:
    """Lee un Excel Prophet y construye un Documento con secciones pre-pobladas."""

    def __init__(self, repo: DocumentoRepository, llm: LLMClient | None = None) -> None:
        self.repo = repo
        self.llm = llm

    def ejecutar(
        self,
        xlsx_bytes: bytes,
        fila_idx: int,
        nombre_modelo: str,
        user_id: str = "default",
    ) -> ResultadoImportacionProphet:
        """Importa el registro Prophet y devuelve el resultado.

        Args:
            xlsx_bytes: contenido del archivo .xlsx.
            fila_idx: índice de fila (base 0) en Descripcion_General para el modelo.
            nombre_modelo: nombre humano del modelo (p. ej. "VNB").
            user_id: ID del usuario que importa.
        """
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True)
        except Exception as e:
            return ResultadoImportacionProphet(
                documento=None,
                advertencias=[f"No se pudo leer el Excel: {e}"],
            )

        # Normaliza nombres de hoja a minúsculas para lookup tolerante
        hojas: dict[str, list[dict[str, Any]]] = {}
        for nombre in wb.sheetnames:
            hojas[nombre.lower().strip()] = self._hoja_a_dicts(wb[nombre])

        secciones = construir_secciones_vacias_prophet()
        advertencias: list[str] = []
        importadas = 0

        for seccion in secciones:
            cat = por_id_prophet(seccion.id)
            if cat is None:
                continue

            hoja_data = self._encontrar_hoja(hojas, seccion.id)
            if not hoja_data:
                advertencias.append(
                    f"Sección '{seccion.nombre}': no se encontró hoja correspondiente en el Excel."
                )
                continue

            contenido = self._extraer(hoja_data, cat, nombre_modelo, fila_idx)
            if contenido:
                seccion.contenido = contenido
                seccion.completitud = "completa"
                importadas += 1
            else:
                advertencias.append(
                    f"Sección '{seccion.nombre}': sin datos extraíbles del Excel."
                )

        doc = Documento(
            user_id=user_id,
            tipo="prophet",
            metadata_modelo=MetadataModelo(nombre_modelo=nombre_modelo),
            secciones=secciones,
        )
        doc.registrar_evento(
            EventoAuditoria(
                timestamp=datetime.now(UTC),
                actor=user_id,
                tipo="documento_creado",
                descripcion=f"Ficha Prophet importada desde Excel: {nombre_modelo}",
                # metadata es dict[str, str] — convertir ints a str
                metadata={
                    "fila_idx": str(fila_idx),
                    "secciones_importadas": str(importadas),
                },
            )
        )
        self.repo.guardar(doc)

        return ResultadoImportacionProphet(
            documento=doc,
            secciones_importadas=importadas,
            secciones_vacias=len(secciones) - importadas,
            advertencias=advertencias,
        )

    # ------------------------------------------------------------------
    # Helpers privados
    # ------------------------------------------------------------------

    def _hoja_a_dicts(self, ws: Any) -> list[dict[str, Any]]:
        """Convierte una hoja Excel a lista de dicts usando la primera fila como headers."""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [
            str(h).strip() if h is not None else f"col_{i}"
            for i, h in enumerate(rows[0])
        ]
        return [
            {
                headers[i]: (str(c).strip() if c is not None else "")
                for i, c in enumerate(row)
            }
            for row in rows[1:]
            if any(c is not None for c in row)
        ]

    def _encontrar_hoja(
        self, hojas: dict[str, list[Any]], seccion_id: str
    ) -> list[dict[str, Any]]:
        """Busca la hoja más apropiada para la sección dada."""
        candidatos = _HOJA_POR_SECCION.get(seccion_id, [])
        for nombre in candidatos:
            if nombre in hojas:
                return hojas[nombre]
        return []

    def _extraer(
        self,
        datos: list[dict[str, Any]],
        cat: SeccionCatalogoProphet,
        nombre_modelo: str,
        fila_idx: int,
    ) -> str | None:
        """Despacha a extracción LLM o heurística según disponibilidad del cliente."""
        if self.llm is not None:
            return self._extraer_con_llm(datos, cat, nombre_modelo)
        return self._extraer_heuristico(datos, cat, nombre_modelo, fila_idx)

    def _extraer_con_llm(
        self,
        datos: list[dict[str, Any]],
        cat: SeccionCatalogoProphet,
        nombre_modelo: str,
    ) -> str | None:
        """Usa Haiku 4.5 para extracción estructurada; cae a heurística si falla."""
        prompt = construir_prompt_extraccion(
            nombre_modelo=nombre_modelo,
            nombre_seccion=cat.nombre,
            schema_campos=list(cat.schema_tabla) if cat.tipo_contenido == "tabla" else [],
            datos_crudos=datos,
            tipo_contenido=cat.tipo_contenido,
        )
        try:
            resp = self.llm.chat(  # type: ignore[union-attr]
                tarea="extraction",
                system_blocks=[TextBlockParam(type="text", text=EXTRAER_SECCION_PROPHET_SYSTEM)],
                messages=[{"role": "user", "content": prompt}],
                max_tokens=4096,
            )
            json.loads(resp.text)  # valida que el LLM devolvió JSON válido
            return resp.text
        except Exception:
            return self._extraer_heuristico(datos, cat, nombre_modelo, 0)

    def _extraer_heuristico(
        self,
        datos: list[dict[str, Any]],
        cat: SeccionCatalogoProphet,
        nombre_modelo: str,
        fila_idx: int,
    ) -> str | None:
        """Extracción sin LLM: mapea datos crudos al formato esperado por sección."""
        if not datos:
            return None

        if cat.tipo_contenido == "campos" and fila_idx < len(datos):
            fila = datos[fila_idx]
            return json.dumps(
                {"contenido": json.dumps(fila, ensure_ascii=False), "advertencias": []},
                ensure_ascii=False,
            )

        if cat.tipo_contenido == "tabla":
            return json.dumps(
                {"filas": datos, "advertencias": ["Extracción heurística — revisar datos"]},
                ensure_ascii=False,
            )

        if cat.tipo_contenido == "texto" and fila_idx < len(datos):
            texto = " | ".join(str(v) for v in datos[fila_idx].values() if v)
            if texto:
                return json.dumps(
                    {"contenido": texto, "advertencias": []},
                    ensure_ascii=False,
                )

        return None
