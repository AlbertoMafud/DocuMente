from __future__ import annotations

import io
import json
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from anthropic.types import TextBlockParam

from src.llm import LLMClient
from src.llm.prompts.extraer_seccion_prophet import EXTRAER_SECCION_PROPHET_SYSTEM


@dataclass
class ModeloProphetInfo:
    fila_idx: int
    nombre: str
    encargado: str
    area: str = ""
    proceso: str = ""


@dataclass
class ResultadoDeteccion:
    modelos: list[ModeloProphetInfo] = field(default_factory=list)
    advertencias: list[str] = field(default_factory=list)


_NOMBRES_HOJA_CATALOGO = {
    "descripcion_general",
    "descripción_general",
    "modelos",
    "catalogo",
    "catálogo",
}
_COLS_NOMBRE = {"proceso", "modelo", "nombre", "nombre_modelo", "nombre del modelo"}
_COLS_ENCARGADO = {"encargado", "responsable", "dueño", "owner", "dueño del modelo"}
_COLS_AREA = {"area", "área"}


class DetectarModelosProphet:
    def __init__(self, llm: LLMClient | None = None) -> None:
        self.llm = llm

    def ejecutar(self, xlsx_bytes: bytes) -> ResultadoDeteccion:
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True)
        except Exception as e:
            return ResultadoDeteccion(advertencias=[f"No se pudo leer el Excel: {e}"])

        hoja = self._encontrar_hoja_catalogo(wb)
        if hoja is None:
            return ResultadoDeteccion(advertencias=["No se encontró hoja de catálogo de modelos."])

        rows = self._hoja_a_dicts(hoja)
        if not rows:
            return ResultadoDeteccion(advertencias=["La hoja de catálogo está vacía."])

        if self.llm is not None:
            return self._detectar_con_llm(rows)
        return self._detectar_heuristico(rows)

    def _encontrar_hoja_catalogo(self, wb: openpyxl.Workbook) -> Any:
        for nombre in wb.sheetnames:
            if nombre.lower().replace(" ", "_") in _NOMBRES_HOJA_CATALOGO:
                return wb[nombre]
        return wb.active

    def _hoja_a_dicts(self, ws: Any) -> list[dict[str, Any]]:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        return [
            {headers[i]: (str(c).strip() if c is not None else "") for i, c in enumerate(row)}
            for row in rows[1:]
            if any(c is not None for c in row)
        ]

    def _detectar_heuristico(self, rows: list[dict[str, Any]]) -> ResultadoDeteccion:
        h_lower = {k.lower(): k for k in rows[0]} if rows else {}
        col_n = next((h_lower[k] for k in _COLS_NOMBRE if k in h_lower), None)
        col_e = next((h_lower[k] for k in _COLS_ENCARGADO if k in h_lower), None)
        col_a = next((h_lower[k] for k in _COLS_AREA if k in h_lower), None)

        modelos = []
        for idx, row in enumerate(rows):
            nombre = (row.get(col_n, "") if col_n else next(iter(row.values()), "")).strip()
            if nombre:
                modelos.append(
                    ModeloProphetInfo(
                        fila_idx=idx,
                        nombre=nombre,
                        encargado=(row.get(col_e, "") if col_e else "").strip(),
                        area=(row.get(col_a, "") if col_a else "").strip(),
                    )
                )
        advertencias = [] if modelos else ["No se encontraron filas de modelos."]
        return ResultadoDeteccion(modelos=modelos, advertencias=advertencias)

    def _detectar_con_llm(self, rows: list[dict[str, Any]]) -> ResultadoDeteccion:
        prompt = (
            f"Datos crudos de catálogo de modelos actuariales Prophet:\n"
            f"```json\n{json.dumps(rows[:20], ensure_ascii=False, indent=2)}\n```\n\n"
            "Identifica todas las filas que representan modelos distintos.\n"
            "Devuelve SOLO este JSON:\n"
            '{"modelos": [{"fila_idx": 0, "nombre": "...", "encargado": "...", "area": ""}], "advertencias": []}'
        )
        try:
            resp = self.llm.chat(  # type: ignore[union-attr]
                tarea="extraction",
                system_blocks=[TextBlockParam(type="text", text=EXTRAER_SECCION_PROPHET_SYSTEM)],
                messages=[{"role": "user", "content": prompt}],
                max_tokens=1024,
            )
            data = json.loads(resp.text)
            modelos = [
                ModeloProphetInfo(
                    fila_idx=m["fila_idx"],
                    nombre=m.get("nombre", ""),
                    encargado=m.get("encargado", ""),
                    area=m.get("area", ""),
                )
                for m in data.get("modelos", [])
                if m.get("nombre", "").strip()
            ]
            return ResultadoDeteccion(modelos=modelos, advertencias=data.get("advertencias", []))
        except Exception as e:
            resultado = self._detectar_heuristico(rows)
            resultado.advertencias.append(f"LLM falló ({e}), se usó detección heurística.")
            return resultado
