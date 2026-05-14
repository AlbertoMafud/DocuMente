"""Lector de XLSX — serializa hojas a texto plano con headers."""

from __future__ import annotations

from typing import IO

from openpyxl import load_workbook


def leer_xlsx(archivo: IO[bytes]) -> str:
    """Devuelve cada hoja como un bloque de texto: nombre + filas tab-separadas.

    Usa `data_only=True` para que las celdas con fórmulas devuelvan su valor
    calculado (no la fórmula literal). Solo procesa las primeras 200 filas y
    20 columnas por hoja para evitar contextos LLM inmanejables.
    """
    archivo.seek(0)
    wb = load_workbook(archivo, data_only=True, read_only=True)

    bloques: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        bloques.append(f"=== Hoja: {sheet_name} ===")
        filas_emitidas = 0
        for fila in ws.iter_rows(values_only=True):
            celdas = [str(c).strip()[:200] if c is not None else "" for c in list(fila)[:20]]
            # Skip filas completamente vacías
            if not any(c.strip() for c in celdas):
                continue
            bloques.append("\t".join(celdas))
            filas_emitidas += 1
            if filas_emitidas >= 200:
                bloques.append("[... contenido truncado a 200 filas ...]")
                break

    wb.close()
    return "\n".join(bloques)
