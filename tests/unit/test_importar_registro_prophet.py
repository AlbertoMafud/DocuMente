from __future__ import annotations

import io
import os

import openpyxl
import pytest

from src.core.usecases.importar_registro_prophet import ImportarRegistroProphet
from src.storage.repositories import DocumentoRepository


def _excel_prophet_minimo() -> bytes:
    """Excel con las 4 hojas del formato Prophet, datos mínimos."""
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Descripcion_General"
    ws1.append(["Area", "Proceso", "Encargado", "Descripcion", "Frecuencia de actualización", "Corridas", "Qué problema ataca"])
    ws1.append(["Rentabilidad", "VNB", "Francisco Carmona", "Modelo de valor nuevo de negocio", "Trimestral", "33,34,36", "Medir rentabilidad"])
    ws1.append(["Rentabilidad", "IRR", "Cynthia Flores", "Internal rate of return", "Trimestral", "33,34", "Medir tasa interna"])

    ws2 = wb.create_sheet("Detalle Runs")
    ws2.append(["# corrida", "Detalle", "Corrida Precedente", "Es ALM?", "Tiempo de ejecución", "Outputs Principales", "Responsable"])
    ws2.append(["33", "IL UDI y USD", "", "No", "45 min", "VNB, Profit", "Carmona"])
    ws2.append(["34", "GMM Individual", "33", "Sí", "90 min", "IRR", "Carmona"])

    ws3 = wb.create_sheet("Variables criticas")
    ws3.append(["Corrida", "Nombre", "Descripción", "Fórmula", "Frecuencia de actualización", "Responsable de la info", "Variables dependientes"])
    ws3.append(["33", "PROF_SOLVM", "Solvency margin profit", "PREM_INC - DEATH_OUTGO", "Trimestral", "Carmona", "PREM_INC,DEATH_OUTGO"])

    ws4 = wb.create_sheet("Conocimiento_Tecnico")
    ws4.append(["Persona", "Ejecutar corridas base", "Modificación de código"])
    ws4.append(["Francisco Carmona", "AVANZADO", "INTERMEDIO"])
    ws4.append(["Cynthia Flores", "INTERMEDIO", "BÁSICO"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture()
def repo(tmp_path):
    os.environ["DATABASE_URL"] = f"sqlite:///{tmp_path}/test.db"
    yield DocumentoRepository()
    # Limpiar para no contaminar otros tests
    os.environ.pop("DATABASE_URL", None)


def test_importa_documento_tipo_prophet(repo) -> None:
    xlsx = _excel_prophet_minimo()
    uc = ImportarRegistroProphet(repo=repo, llm=None)
    resultado = uc.ejecutar(xlsx_bytes=xlsx, fila_idx=0, nombre_modelo="VNB")
    assert resultado.documento is not None
    assert resultado.documento.tipo == "prophet"
    assert resultado.documento.metadata_modelo.nombre_modelo == "VNB"


def test_secciones_identificacion_pre_poblada(repo) -> None:
    xlsx = _excel_prophet_minimo()
    uc = ImportarRegistroProphet(repo=repo, llm=None)
    resultado = uc.ejecutar(xlsx_bytes=xlsx, fila_idx=0, nombre_modelo="VNB")
    assert resultado.documento is not None
    doc = resultado.documento
    assert doc.seccion_por_id("identificacion") is not None


def test_excel_invalido_no_lanza_excepcion(repo) -> None:
    uc = ImportarRegistroProphet(repo=repo, llm=None)
    resultado = uc.ejecutar(xlsx_bytes=b"no es excel", fila_idx=0, nombre_modelo="VNB")
    assert resultado.documento is None or len(resultado.advertencias) > 0


def test_documento_persiste_en_repo(repo) -> None:
    xlsx = _excel_prophet_minimo()
    uc = ImportarRegistroProphet(repo=repo, llm=None)
    resultado = uc.ejecutar(xlsx_bytes=xlsx, fila_idx=0, nombre_modelo="VNB")
    if resultado.documento:
        doc_repo = repo.obtener(resultado.documento.id)
        assert doc_repo is not None
        assert doc_repo.tipo == "prophet"


def test_columna_faltante_no_rompe_import(repo) -> None:
    """Excel sin hoja Variables criticas — el import completa lo que puede."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Descripcion_General"
    ws.append(["Proceso", "Encargado"])
    ws.append(["VNB", "Carmona"])
    buf = io.BytesIO()
    wb.save(buf)

    uc = ImportarRegistroProphet(repo=repo, llm=None)
    resultado = uc.ejecutar(xlsx_bytes=buf.getvalue(), fila_idx=0, nombre_modelo="VNB")
    assert resultado is not None
